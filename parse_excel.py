#!/usr/bin/env python3
"""
parse_excel.py — OT Monitor 2026 data pipeline
================================================

Reads the "monitoring" sheet of the attendance/overtime workbook and produces
a compact JSON file (dashboard_data.json) that the dashboard (docs/index.html)
fetches at load time.

WHY THIS SCRIPT EXISTS
-----------------------
The dashboard used to require someone to manually open the HTML file and
click "Upload Excel" — which only updates OT data in that one person's
browser. This script moves that step into the repo/CI pipeline so that:

  1. A manager updates the Excel file in `data/`.
  2. GitHub Actions runs this script automatically.
  3. It writes `docs/data/dashboard_data.json`.
  4. GitHub Pages serves the updated JSON to EVERY viewer of the same link.

No more "my dashboard shows different numbers than yours" — everyone reads
the same committed JSON file.

SHEET LAYOUT THIS SCRIPT EXPECTS ("monitoring" sheet)
------------------------------------------------------
  Row  = header row (auto-detected: the row containing "Full Name")
  Row+1 = a "day-of-week" row (formula results like Mon/Tue/.../Sun, or
          "Chính thức" / "Làm thêm" for the two summary columns of each
          month block)

  Columns, left to right:
    STT | ID Code | Full Name | Position | Shift |
    [Year summary: Chính thức | Làm thêm] |
    [Month 1 summary: Chính thức | Làm thêm] [Month 1 daily columns...] |
    [Month 2 summary: Chính thức | Làm thêm] [Month 2 daily columns...] |
    ...

  Each employee occupies up to 3 consecutive rows identified by the value
  in the "Shift" column:
    D-shift   -> day shift attendance (ignored for OT purposes)
    N-shift   -> night shift attendance (ignored for OT purposes)
    Overtime  -> the row this script actually reads OT hours from

  The month-block "Chính thức"/"Làm thêm" column pair is a *merged* header
  whose displayed date can coincidentally equal the first daily column's
  date. To avoid double-counting / misreading, this script uses the
  DAY-OF-WEEK label row (Mon/Tue/.../Sun) — not the date-typed header row —
  to positively identify which columns are real daily columns. This fixes a
  bug present in the browser-side parser embedded in the dashboard, where a
  month's total could silently overwrite the first day's value.

USAGE
-----
  python parse_excel.py <input.xlsx> <output.json> [--sheet monitoring]

  Exits non-zero (and prints a clear error) if the workbook doesn't look
  like the expected template, so a bad upload fails the CI job loudly
  instead of silently publishing empty/wrong data.
"""
import sys
import json
import argparse
import datetime as dt
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: run `pip install openpyxl` first.")

MONTH_TARGET = 40      # h / person / month
WEEK_TARGET = 40        # h / person / week
DAY_TARGET = 2           # h / person / day
YEAR_TARGET = 300       # h / person / year (informational only)

WEEKDAY_LABELS = {"Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"}
SUMMARY_REGULAR_LABELS = {"chính thức", "chinh thuc"}
SUMMARY_OT_LABELS = {"làm thêm", "lam them"}


class SheetFormatError(Exception):
    pass


def parse_num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-", "—", "x", "X"):
        return 0.0
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def norm(v):
    return str(v).strip().lower() if v is not None else ""


def find_sheet(wb, requested=None):
    if requested:
        for name in wb.sheetnames:
            if name.strip().lower() == requested.strip().lower():
                return wb[name]
        raise SheetFormatError(f'Sheet "{requested}" not found. Available: {wb.sheetnames}')
    for name in wb.sheetnames:
        if name.strip().lower() == "monitoring":
            return wb[name]
    for name in wb.sheetnames:
        if "monitoring" in name.strip().lower():
            return wb[name]
    raise SheetFormatError(f'No "monitoring" sheet found. Available: {wb.sheetnames}')


def find_header_row(ws, max_scan_rows=40, max_scan_cols=40):
    for r in range(1, max_scan_rows + 1):
        for c in range(1, max_scan_cols + 1):
            v = ws.cell(row=r, column=c).value
            if v and "full name" in str(v).lower():
                return r, c
    raise SheetFormatError('Could not locate a "Full Name" header cell in the first 40 rows/cols.')


def build_column_map(ws, header_row, name_col, max_col):
    """Locate STT / ID / Name / Position / Shift columns near the Full Name column,
    and classify every later column as either a daily-OT column or a month
    summary column, using the day-of-week label row (header_row + 1)."""
    day_row = header_row + 1

    def hdr(c):
        return norm(ws.cell(row=header_row, column=c).value)

    col_id = col_pos = col_shift = None
    for c in range(max(1, name_col - 4), name_col):
        h = hdr(c)
        if "id" in h:
            col_id = c
    for c in range(name_col + 1, name_col + 6):
        h = hdr(c)
        if "position" in h:
            col_pos = c
        elif h == "shift":
            col_shift = c
            break
    if col_id is None or col_shift is None:
        raise SheetFormatError("Could not locate ID Code / Shift columns next to Full Name.")

    daily_cols = {}     # excel_col -> datetime.date
    month_ot_col = {}   # month_number(1-12) -> excel_col holding that month's OT total

    c = col_shift + 1
    seen_pairs = 0
    while c <= max_col:
        lbl = norm(ws.cell(row=day_row, column=c).value)
        if lbl in SUMMARY_REGULAR_LABELS:
            nxt = norm(ws.cell(row=day_row, column=c + 1).value)
            if nxt in SUMMARY_OT_LABELS:
                seen_pairs += 1
                if seen_pairs > 1:
                    # first pair = year-to-date summary block, skip it;
                    # subsequent pairs = per-month summary blocks
                    month_date = ws.cell(row=header_row, column=c + 2).value
                    if isinstance(month_date, dt.datetime):
                        month_date = month_date.date()
                    if isinstance(month_date, dt.date):
                        month_ot_col[month_date.month] = c + 1
                c += 2
                continue
        elif ws.cell(row=day_row, column=c).value in WEEKDAY_LABELS:
            hval = ws.cell(row=header_row, column=c).value
            if isinstance(hval, dt.datetime):
                hval = hval.date()
            if isinstance(hval, dt.date):
                daily_cols[c] = hval
        c += 1

    if not daily_cols:
        raise SheetFormatError("Found 0 daily OT columns — sheet layout may have changed.")

    return {
        "col_id": col_id, "col_name": name_col, "col_pos": col_pos, "col_shift": col_shift,
        "daily_cols": daily_cols, "month_ot_col": month_ot_col,
    }


def iso_week(d):
    return d.isocalendar()[1]


def extract_employees(ws, cmap, header_row, max_row):
    employees = []
    r = header_row + 2  # first data row
    n_cols = max(cmap["daily_cols"].keys() | cmap["month_ot_col"].values(), default=cmap["col_shift"])

    while r <= max_row:
        shift_val = norm(ws.cell(row=r, column=cmap["col_shift"]).value)
        if shift_val == "d-shift":
            emp_id = ws.cell(row=r, column=cmap["col_id"]).value
            emp_id = str(emp_id).strip() if emp_id is not None else ""
            name = ws.cell(row=r, column=cmap["col_name"]).value
            name = str(name).strip() if name is not None else ""
            position = ws.cell(row=r, column=cmap["col_pos"]).value if cmap["col_pos"] else None
            position = str(position).strip() if position is not None else ""

            ot_row = None
            for off in range(1, 4):
                if r + off > max_row:
                    break
                s = norm(ws.cell(row=r + off, column=cmap["col_shift"]).value)
                if "overtime" in s:
                    ot_row = r + off
                    break
                if off >= 2 and s == "d-shift":
                    break  # ran into next employee without finding an Overtime row

            monthly_ots = [0.0] * 12
            weekday_ot_by_month = [0.0] * 12
            weekend_ot_by_month = [0.0] * 12
            weekly_ot, weekly_benefit, daily_ot = {}, {}, {}

            if ot_row and emp_id:
                for col, d in cmap["daily_cols"].items():
                    v = parse_num(ws.cell(row=ot_row, column=col).value)
                    if v == 0:
                        continue
                    mi = d.month - 1
                    dow = d.weekday()  # 0=Mon .. 6=Sun
                    if dow >= 5:
                        weekend_ot_by_month[mi] += v
                    else:
                        weekday_ot_by_month[mi] += v
                    wk = iso_week(d)
                    weekly_ot[wk] = round(weekly_ot.get(wk, 0.0) + v, 2)
                    if dow >= 5:
                        weekly_benefit[wk] = round(weekly_benefit.get(wk, 0.0) + v, 2)
                    key = d.strftime("%Y-%m-%d")
                    daily_ot[key] = round(daily_ot.get(key, 0.0) + v, 2)

                for month_num, col in cmap["month_ot_col"].items():
                    v = parse_num(ws.cell(row=ot_row, column=col).value)
                    if v > 0:
                        monthly_ots[month_num - 1] = round(v, 2)

                # fallback: derive monthly total from daily sums if the summary
                # cell was blank/zero but daily data exists for that month
                for mi in range(12):
                    if monthly_ots[mi] == 0:
                        s = weekday_ot_by_month[mi] + weekend_ot_by_month[mi]
                        if s > 0:
                            monthly_ots[mi] = round(s, 2)

            if emp_id:
                year_ot = round(sum(monthly_ots), 2)
                employees.append({
                    "id": emp_id,
                    "name": name,
                    "position": position,
                    "year_ot": year_ot,
                    "monthly_ots": [round(v, 2) for v in monthly_ots],
                    "weekday_ot_by_month": [round(v, 2) for v in weekday_ot_by_month],
                    "weekend_ot_by_month": [round(v, 2) for v in weekend_ot_by_month],
                    "weekly_ot": {str(k): v for k, v in sorted(weekly_ot.items())},
                    "weekly_benefit": {str(k): v for k, v in sorted(weekly_benefit.items())},
                    "daily_ot": daily_ot,
                })
        r += 1
    return employees


def parse_workbook(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = find_sheet(wb, sheet_name)
    header_row, name_col = find_header_row(ws)
    cmap = build_column_map(ws, header_row, name_col, ws.max_column)
    employees = extract_employees(ws, cmap, header_row, ws.max_row)
    if not employees:
        raise SheetFormatError("Parsed 0 employees — check that Shift column contains 'D-shift' rows.")
    return employees


def main():
    ap = argparse.ArgumentParser(description="Parse the OT monitoring workbook into dashboard_data.json")
    ap.add_argument("input", help="Path to the .xlsx workbook")
    ap.add_argument("output", help="Path to write dashboard_data.json")
    ap.add_argument("--sheet", default=None, help='Sheet name (default: auto-detect "monitoring")')
    args = ap.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        sys.exit(f"Input file not found: {in_path}")

    try:
        employees = parse_workbook(in_path, args.sheet)
    except SheetFormatError as e:
        sys.exit(f"ERROR parsing '{in_path.name}': {e}")

    out = {
        "employees": employees,
        "updatedAt": dt.datetime.now(dt.timezone.utc).isoformat(),
        "fileName": in_path.name,
        "targets": {
            "monthly": MONTH_TARGET, "weekly": WEEK_TARGET,
            "daily": DAY_TARGET, "yearly": YEAR_TARGET,
        },
    }

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=None, separators=(",", ":")), encoding="utf-8")

    total_year_ot = sum(e["year_ot"] for e in employees)
    print(f"Parsed {len(employees)} employees, total year OT = {total_year_ot:.1f}h")
    print(f"Wrote {out_path} ({out_path.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
